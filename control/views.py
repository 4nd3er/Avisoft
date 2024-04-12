import json
from django.shortcuts import render, HttpResponse, redirect
from django.shortcuts import get_object_or_404
from .models import *
from .forms import *
from django.core.serializers import serialize
from django.urls import reverse_lazy
from django.db.models import Q
from django.views.generic import TemplateView, ListView, UpdateView, CreateView, DeleteView, ListView
from django.contrib.auth import login, logout, authenticate
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_http_methods
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from django.views.decorators.http import require_http_methods
from django.core.exceptions import ObjectDoesNotExist
import datetime
from datetime import timedelta
import time

def is_ajax(request):
    return request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest'

def fechaActual():
    todays_Date = datetime.date.fromtimestamp(time.time())
    dateCurrent = todays_Date.isoformat()
    return dateCurrent

@require_http_methods(['GET'])
def GalponData(request, id):
    dataGalpon = Galpones.objects.all().values().get(id=id)
    dataProd = ProduccionDiaria.objects.filter(id_galpon=id, fecha=fechaActual()).values()
    totalHuevos = 0
    for dato in dataProd:
        totalHuevos += int(dato['cantidad'])
    response = JsonResponse({'dataGalpon': dataGalpon, 'dataProd': totalHuevos}, safe=False)
    return response

@require_http_methods(['GET'])
def GalponDataDes(request, id):
    dataGalpon = Galpones.objects.all().values().get(id=id)
    return JsonResponse(dataGalpon, safe=False)

# ! Modulo de inicio e interfaces
def registrarse(request):
    registro = UsuarioForm()
    if request.method == 'POST':
        registro = UsuarioForm(request.POST, request.FILES)
        if registro.is_valid():
            registro.is_active = False
            registro.save()
            messages.success(request, f'Te has registrado exitosamente {registro.nombre}')
            return redirect('inicio')
    return render(request, 'inicio_sesion/registrarse.html', { 'form': registro })


def inicio(request):
    if request.user.is_authenticated:
        return redirect('interfaz')
    else:
        if request.method == 'POST':
            documento = request.POST.get('documento')
            password = request.POST.get('password')

            if documento == "" and password == "":
                messages.warning(request, 'Digita en los campos correspondientes para el inicio de sesion')
                return render(request, 'inicio_sesion/inicio.html')
            user = authenticate(request, documento = documento, password = password)

            userFilter = Usuario.objects.filter(documento=documento).values_list('is_active', flat=True)
            if user is not None:
                login(request, user)
                return redirect('interfaz')
            elif userFilter:
                if not userFilter[0]:
                    messages.error(request, 'Usuario no activado, comunicate con el administrador para más información')
                else:
                    messages.error(request, 'Numero de documento y/o contraseña incorrectos, vuelve a intentarlo')
            else:
                messages.error(request, 'Numero de documento y/o contraseña incorrectos, vuelve a intentarlo')
    return render(request, 'inicio_sesion/inicio.html')

def acerca_de(request):
    # Lógica si es necesario
    acercaDe = True
    return render(request, 'usuarios/acercaDe/acerca_de.html', { 'acercaDe': acercaDe })

class contrasena(LoginRequiredMixin, ListView):
    template_name = 'usuarios/cambioPswrd/password.html'
    form_class = cambioPasswordForm
    success_url = reverse_lazy('interfaz')

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name, {'form': self.form_class})

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST)
        if form.is_valid():
            user = Usuario.objects.filter(id = request.user.id)
            if user.exists():
                user = user.first()
                user.set_password(form.cleaned_data.get('password1'))
                user.save()
                logout(request)
                return redirect(self.success_url)
            return redirect(self.success_url)
        else:
            form = self.form_class(request.POST)
            messages.error(request, 'Las contraseñas no coinciden')
            return render(request, self.template_name, {'form': form})


def interfaz(request):
    return render(request, 'interfaz/interfaces.html')


def logout_usuario(request):
    logout(request)
    return redirect('inicio')
# ! Modulo de inicio e interfaces


# ! Modulo de registro diario
class registroDiario(ListView):
    template_name = 'registro_diario/registro_diario.html'
    model = Registrodiario

    def get_queryset(self):
        select = self.request.GET.get('search')
        busqueda = self.request.GET.get('buscar')
        busquedaDate = self.request.GET.get('buscarDate')
        if select == 'date':
            query = self.model.objects.filter(
                Q(fecha__icontains = busquedaDate)
                ).distinct().order_by('-id')
            busqueda = ''
        elif select == 'input':
            query = self.model.objects.filter(
                Q(id_galpon__nombre_galpon__icontains = busqueda) |
                Q(id_gallinas__id_linea__nombre__icontains = busqueda) |
                Q(id_gallinas__cantidad_gallinas__icontains = busqueda) |
                Q(id_gallinas__procedencia__icontains = busqueda) |
                Q(id_alimentacion__kg_total__icontains = busqueda) |
                Q(id_alimentacion__gr_gallina_dia__icontains = busqueda) |
                Q(id_alimentacion__id_tipo_alimento__nombre__icontains = busqueda) |
                Q(id_producciondiaria__id_jornada__jornada__icontains = busqueda) |
                Q(id_producciondiaria__id_tipo_huevo__tipos_huevos__icontains = busqueda) |
                Q(id_producciondiaria__id_usuario__nombre__icontains = busqueda) |
                Q(id_mortades__id_tipo_descarte__tipo__icontains = busqueda)
                ).distinct()
            busquedaDate = ''
        else:
            query = self.model.objects.filter(fecha=fechaActual()).order_by('-id')
        return query

    def get_context_data(self, **kwargs):
        contexto = {}
        contexto['registroDiario'] = self.get_queryset()
        return contexto

    def get(self, request, *args, **kwargs):
        registroDiarioMenu = True
        return render(request, self.template_name, {'registroDiario': self.get_queryset(), 'registroDiarioMenu': registroDiarioMenu})
    
    def post(self, request, *args, **kwargs):
        registroDiarioMenu = True
        query = self.get_queryset()
        if not query:
            messages.error(request, 'Debes buscar algun dato o registrar en los modulos correspondientes para generar el reporte')
            return render(request, self.template_name, {'registroDiario': self.get_queryset(), 'registroDiarioMenu': registroDiarioMenu})
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Hoja1'

            ws['B2'].alignment = Alignment(horizontal = 'center', vertical = 'center')
            ws['B2'].border = Border(left = Side(border_style = 'thin'), right = Side(border_style = 'thin'),
                                        top = Side(border_style = 'thin'), bottom = Side(border_style = 'thin'))
            ws['B2'].fill = PatternFill(start_color = '39A900', fill_type = 'solid')
            ws['B2'].font = Font(name = 'Arial', size = 15, bold = True, color = 'FFFFFF')
            ws['B2'] = f'REPORTE {self.model.__str__().upper()}'

            ws.merge_cells('B2:G2')
            listColumn = ['B', 'C', 'D', 'E', 'F','G']
            listName = ['Galpon', 'Lote de Gallinas', 'Alimentación', 'Proucción diaria', 'Mortalidad y descarte', 'Fecha']
            countName = 0
            count = 3
            ws.row_dimensions[2].height = 25
            for i in listColumn:
                ws.column_dimensions[i].width = 55
                ws[f'{listColumn[countName]}3'].alignment = Alignment(horizontal = 'center', vertical = 'center')
                ws[f'{listColumn[countName]}3'].border = Border(left = Side(border_style = 'thin'), right = Side(border_style = 'thin'),
                                            top = Side(border_style = 'thin'), bottom = Side(border_style = 'thin'))
                ws[f'{listColumn[countName]}3'].fill = PatternFill(start_color = 'FFCE40', fill_type = 'solid')
                ws[f'{listColumn[countName]}3'].font = Font(name = 'Arial', size = 11)
                ws[f'{listColumn[countName]}3'] = listName[countName]
                count += 1
                countName += 1
            
            # Pintamos los datos en el reporte
            listName = ['id_galpon', 'id_gallinas', 'id_alimentacion', 'id_producciondiaria', 'id_mortades', 'fecha']
            countColumn = 2
            for i in listName:
                countRow = 4
                for q in query:
                    ws.cell(row = countRow, column = countColumn).alignment = Alignment(horizontal = 'center', vertical = 'center')
                    ws.cell(row = countRow, column = countColumn).border = Border(left = Side(border_style = 'thin'), right = Side(border_style = 'thin'),
                                                    top = Side(border_style = 'thin'), bottom = Side(border_style = 'thin'))
                    ws.cell(row = countRow, column = countColumn).fill = PatternFill(start_color = 'FBFBE2', fill_type = 'solid')
                    ws.cell(row = countRow, column = countColumn).font = Font(name = 'Arial', size = '11')
                    if i == 'id_galpon':
                        if hasattr(q, 'id_galpon'):
                            valueRow = q.id_galpon.__str__()
                        else:
                            valueRow = 'No aplica'
                    elif i == 'id_gallinas':
                        if hasattr(q, 'id_gallinas'):
                            valueRow = q.id_gallinas.__str__()
                        else:
                            valueRow = 'No aplica'
                    elif i == 'id_alimentacion':
                        if hasattr(q, 'id_alimentacion'):
                            valueRow = q.id_alimentacion.__str__()
                        else:
                            valueRow = 'No aplica'
                    elif i == 'id_producciondiaria':
                        if hasattr(q, 'id_producciondiaria'):
                            valueRow = q.id_producciondiaria.__str__()
                        else:
                            valueRow = 'No aplica'
                    elif i == 'id_mortades':
                        if hasattr(q, 'id_mortades'):
                            cant_muertas = getattr(q.id_mortades, 'cant_muertas', 'cant_muertas')
                            cant_descarte = getattr(q.id_mortades, 'cant_descarte', 'cant_descarte')
                            valueRow = int(cant_muertas) + int(cant_descarte)
                        else:
                            valueRow = 'No aplica'
                    elif i == 'fecha':
                        valueRow = str(getattr(q, i))
                    else:
                        valueRow = getattr(q, i)
                    ws.cell(row=countRow, column=countColumn).value = valueRow
                    countRow += 1
                countColumn += 1

            # Nombre del archivo
            nombreArchivo = f'REPORTE {self.model.__str__().upper()}.xlsx'
            # Definir el tipo de respuesta
            response = HttpResponse(content_type = 'application/ms-excel')
            contenido = "attachment; filename = {0}".format(nombreArchivo)
            response['Content-Disposition'] = contenido
            wb.save(response)
            return response
        return render(request, self.template_name, {'registroDiario': self.get_queryset(), 'registroDiarioMenu': registroDiarioMenu})
# ! Modulo de registro diario


# ! Modulo de alimentacion
class Alimentacionn(ListView):
    model = Alimentacion
    template_name = 'alimentacion/alimentacion.html'
    
    def get_queryset(self):
        select = self.request.GET.get('search')
        busqueda = self.request.GET.get('buscar')
        busquedaDate = self.request.GET.get('buscarDate')
        if select == 'date':
            query = self.model.objects.filter(
                Q(fecha__icontains = busquedaDate)
                ).distinct().order_by('-id')
            busqueda = ''
        elif select == 'input':
            query = self.model.objects.filter(
                Q(fecha__icontains = busqueda) |
                Q(id_galpon__nombre_galpon__icontains = busqueda) |
                Q(gr_gallina_dia__icontains = busqueda) |
                Q(kg_total__icontains = busqueda) |
                Q(bultos_total__icontains = busqueda) |
                Q(c_a__icontains = busqueda) |
                Q(id_tipo_alimento__nombre__icontains = busqueda)
            ).distinct().order_by('-id')
            busquedaDate = ''
        else:
            return self.model.objects.filter(fecha=fechaActual()).order_by('-id')
        return query

    def get_context_data(self, **kwargs):
        contexto = {}
        contexto["alimentacion"] = self.get_queryset()
        return contexto

    def get(self, request, *args, **kwargs):
        aliMenu = True
        if is_ajax(request=request):
            return HttpResponse(serialize('json', self.get_context_data()), 'application/json')
        else:
            return render(request, self.template_name, {'alimentacion': self.get_queryset(), 'aliMenu': aliMenu})

    def post(self, request, *args, **kwargs):
        query = self.get_queryset()
        if not query:
            messages.error(request, 'Debes buscar algun dato para generar el reporte')
            return render(request, self.template_name, {'alimentacion': self.get_queryset()})
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Hoja1'

            ws['B2'].alignment = Alignment(horizontal = 'center', vertical = 'center')
            ws['B2'].border = Border(left = Side(border_style = 'thin'), right = Side(border_style = 'thin'),
                                        top = Side(border_style = 'thin'), bottom = Side(border_style = 'thin'))
            ws['B2'].fill = PatternFill(start_color = '39A900', fill_type = 'solid')
            ws['B2'].font = Font(name = 'Arial', size = 15, bold = True, color = 'FFFFFF')
            ws['B2'] = f'REPORTE {self.model.__name__.upper()}'

            ws.merge_cells('B2:H2')
            listColumn = ['B', 'C', 'D', 'E', 'F','G', 'H']
            listName = ['Galpon', 'Gr/Gallina/Día', 'Kg Total', 'Bultos Total', 'Conversion Alimenticia', 'Tipo de alimento' ,'Fecha']
            countName = 0
            count = 3
            ws.row_dimensions[2].height = 25
            for i in listColumn:
                ws.column_dimensions[i].width = 55
                ws[f'{listColumn[countName]}3'].alignment = Alignment(horizontal = 'center', vertical = 'center')
                ws[f'{listColumn[countName]}3'].border = Border(left = Side(border_style = 'thin'), right = Side(border_style = 'thin'),
                                            top = Side(border_style = 'thin'), bottom = Side(border_style = 'thin'))
                ws[f'{listColumn[countName]}3'].fill = PatternFill(start_color = 'FFCE40', fill_type = 'solid')
                ws[f'{listColumn[countName]}3'].font = Font(name = 'Arial', size = 11)
                ws[f'{listColumn[countName]}3'] = listName[countName]
                count += 1
                countName += 1
            
            # Pintamos los datos en el reporte
            listName = ['id_galpon', 'gr_gallina_dia', 'kg_total', 'bultos_total', 'c_a', 'id_tipo_alimento' ,'fecha']
            countColumn = 2
            for i in listName:
                countRow = 4
                for q in query:
                    ws.cell(row = countRow, column = countColumn).alignment = Alignment(horizontal = 'center', vertical = 'center')
                    ws.cell(row = countRow, column = countColumn).border = Border(left = Side(border_style = 'thin'), right = Side(border_style = 'thin'),
                                                    top = Side(border_style = 'thin'), bottom = Side(border_style = 'thin'))
                    ws.cell(row = countRow, column = countColumn).fill = PatternFill(start_color = 'FBFBE2', fill_type = 'solid')
                    ws.cell(row = countRow, column = countColumn).font = Font(name = 'Arial', size = '11')
                    if i == 'id_galpon':
                        if hasattr(q, 'id_galpon'):
                            valueRow = q.id_galpon.__str__()
                    elif i == 'id_tipo_alimento':
                        if hasattr(q, 'id_tipo_alimento'):
                            valueRow = q.id_tipo_alimento.__str__()
                    elif i == 'fecha':
                        valueRow = str(getattr(q, i))
                    else:
                        valueRow = getattr(q, i)
                    ws.cell(row=countRow, column=countColumn).value = valueRow
                    countRow += 1
                countColumn += 1

            # Nombre del archivo
            nombreArchivo = f'REPORTE {self.model.__name__.upper()}.xlsx'
            # Definir el tipo de respuesta
            response = HttpResponse(content_type = 'application/ms-excel')
            contenido = "attachment; filename = {0}".format(nombreArchivo)
            response['Content-Disposition'] = contenido
            wb.save(response)
            return response
        return render(request, self.template_name, {'alimentacion': self.get_queryset()})

class crearAlimentacion(CreateView):
    model = Alimentacion
    template_name = 'alimentacion/crear.html'
    form_class = AlimentacionForm
    success_url = reverse_lazy('alimentacion')

    def get(self, request, *args, **kwargs):
        try:
            mortaDes = MortalidadDescarte.objects.filter(fecha=fechaActual())
        except:
            mortaDes = 0
        return render(request, self.template_name, {'form': self.form_class(), 'mortaDes': mortaDes})

    def post(self, request, *args, **kwargs):
        if is_ajax(request=request):
            form = self.form_class(request.POST)
            if form.is_valid():
                galponForm = form.cleaned_data['id_galpon']
                form.save()
                alimentacionSaved = Alimentacion.objects.latest('id')
                try:
                    registroDiarioSaved = Registrodiario.objects.filter(fecha=fechaActual(), id_galpon=galponForm).last()
                except ObjectDoesNotExist:
                    registroDiarioSaved = 0
                id_gallinas = None
                id_producciondiaria = None
                id_mortades = None
                if hasattr(registroDiarioSaved, 'id_gallinas'):
                    id_gallinas = registroDiarioSaved.id_gallinas
                if hasattr(registroDiarioSaved, 'id_producciondiaria'):
                    id_producciondiaria = registroDiarioSaved.id_producciondiaria
                if hasattr(registroDiarioSaved, 'id_mortades'):
                    id_mortades = registroDiarioSaved.id_mortades
                if id_producciondiaria or id_gallinas:
                    registrosDiariosSavedAll = Registrodiario.objects.filter(id_galpon=galponForm, fecha=fechaActual())
                    for registro in registrosDiariosSavedAll:
                        registro.id_alimentacion = alimentacionSaved
                        if hasattr(registroDiarioSaved, 'id_gallinas'):
                            registro.id_gallinas = id_gallinas
                        if hasattr(registroDiarioSaved, 'id_mortades'):
                            registro.id_mortades = registroDiarioSaved.id_mortades
                        registro.save()
                elif registroDiarioSaved:
                    if not hasattr(registroDiarioSaved, 'id_alimentacion') or hasattr(registroDiarioSaved, 'id_alimentacion'):
                        registroDiarioSaved.id_alimentacion = alimentacionSaved
                        registroDiarioSaved.save()
                else:
                    registro = Registrodiario(id_galpon=galponForm ,id_alimentacion=alimentacionSaved)
                    registro.save()
                mensaje = f'{self.model.__name__} registrado correctamente!'
                error = 'no hay error'
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 201
                return response
            else:
                mensaje = f'{self.model.__name__} no se pudo registrar'
                error = form.errors
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 400
                return response
        else:
            return redirect('alimentacion')

class editarAlimentacion(UpdateView):
    model = Alimentacion
    template_name = 'alimentacion/editar.html'
    form_class = AlimentacionForm
    success_url = reverse_lazy('alimentacion')

    def post(self, request, *args, **kwargs):
        if is_ajax(request=request):
            form = self.form_class(request.POST, instance = self.get_object())
            if form.is_valid():
                form.save()
                mensaje = f'{self.model.__name__} actualizado correctamente!'
                error = 'no hay error'
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 201
                return response
            else:
                mensaje = f'{self.model.__name__} no se pudo actualizar'
                error = form.errors
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 400
                return response
        else:
            redirect('alimentacion')

class confirmarEliminarAlimentacion(DeleteView):
    model = Alimentacion
    template_name = 'alimentacion/alimentacion_confirm_delete.html'
    success_url = reverse_lazy('alimentacion')

    def post(self, request, *args, **kwargs):
        return render(request, self.template_name)

def eliminarAlimentacion(request, id):
    eliminar = Alimentacion.objects.get(id = id)
    eliminar.delete()
    return redirect('alimentacion')
# ! Modulo de alimentacion


# ! Modulo de fichas
class Fichass(ListView):
    model = Ficha
    template_name = 'fichas/fichas.html'

    def get_queryset(self):
        select = self.request.GET.get('search')
        busqueda = self.request.GET.get('buscar')
        busquedaDate = self.request.GET.get('buscarDate')
        if select == 'date':
            query = self.model.objects.filter(
                Q(fecha__icontains = busquedaDate)
                ).distinct().order_by('-id')
            busqueda = ''
        elif select == 'input':
            query = self.model.objects.filter(
                Q(fecha_regis__icontains = busqueda) |
                Q(num_ficha__icontains = busqueda) |
                Q(id_nombreficha__nombre__icontains = busqueda) |
                Q(estado_ficha__icontains = busqueda) 
            ).distinct().order_by('-id_ficha')
            busquedaDate = ''
        else:
            query = 0
        return query

    def get_context_data(self, **kwargs):
        contexto = {}
        contexto["fichas"] = self.get_queryset()
        return contexto

    def get(self, request, *args, **kwargs):
        user = Usuario.objects.filter(id = request.user.id).values_list('is_staff', flat = True)
        if user[0] == True:
            if is_ajax(request=request):
                return HttpResponse(serialize('json', self.get_context_data()), 'application/json')
            else:
                return render(request, self.template_name, {'fichas': self.get_queryset()})
        else:
            return redirect('interfaz')

class crearFicha(CreateView):
    model = Ficha
    template_name = 'fichas/crear.html'
    form_class = FichaForm
    success_url = reverse_lazy('fichas')

    def post(self, request, *args, **kwargs):
        if is_ajax(request=request):
            form = self.form_class(request.POST)
            if form.is_valid():
                form.save()
                mensaje = f'{self.model.__name__} registrado correctamente!'
                error = 'no hay error'
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 201
                return response
            else:
                mensaje = f'{self.model.__name__} no se pudo registrar'
                error = form.errors
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 400
                return response
        else:
            return redirect('fichas')

class editarFicha(UpdateView):
    model = Ficha
    template_name = 'fichas/editar.html'
    form_class = FichaForm
    success_url = reverse_lazy('fichas')

    def post(self, request, *args, **kwargs):
        if is_ajax(request=request):
            form = self.form_class(request.POST, instance = self.get_object())
            if form.is_valid():
                form.save()
                mensaje = f'{self.model.__name__} actualizado correctamente!'
                error = 'no hay error'
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 201
                return response
            else:
                mensaje = f'{self.model.__name__} no se pudo actualizar'
                error = form.errors
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 400
                return response
        else:
            redirect('fichas')

class confirmarEliminarFicha(DeleteView):
    model = Ficha
    template_name = 'fichas/fichas_confirm_delete.html'
    success_url = reverse_lazy('fichas')

    def post(self, request, *args, **kwargs):
        return render(request, self.template_name)

def eliminarFicha(request, id_ficha):
    eliminar = Ficha.objects.get(id_ficha = id_ficha)
    eliminar.delete()
    return redirect('fichas')
# ! Modulo de fichas


# ! Modulo de nombre de fichas
class FichaNombres(ListView):
    model = Nombreficha
    template_name = 'nombre_fichas/nombre_fichas.html'

    def get_queryset(self):
        select = self.request.GET.get('search')
        busqueda = self.request.GET.get('buscar')
        busquedaDate = self.request.GET.get('buscarDate')
        if select == 'date':
            query = self.model.objects.filter(
                Q(fecha__icontains = busquedaDate)
                ).distinct().order_by('-id')
            busqueda = ''
        elif select == 'input':
            query = self.model.objects.filter(
                Q(nombre__icontains = busqueda)
            ).distinct().order_by('-id')
            busquedaDate = ''
        else:
            query = 0
        return query

    def get_context_data(self, **kwargs):
        contexto = {}
        contexto["fichas"] = self.get_queryset()
        return contexto

    def get(self, request, *args, **kwargs):
        user = Usuario.objects.filter(id = request.user.id).values_list('is_staff', flat = True)
        if user[0] == True:
            if is_ajax(request=request):
                return HttpResponse(serialize('json', self.get_context_data()), 'application/json')
            else:
                return render(request, self.template_name, {'fichas': self.get_queryset()})
        else:
            return redirect('interfaz')

class crearFichaNombre(CreateView):
    model = Nombreficha
    template_name = 'nombre_fichas/crear.html'
    form_class = NombreFichaForm
    success_url = reverse_lazy('fichasNombres')

    def post(self, request, *args, **kwargs):
        if is_ajax(request=request):
            form = self.form_class(request.POST)
            if form.is_valid():
                form.save()
                mensaje = f'{self.model.__name__} registrado correctamente!'
                error = 'no hay error'
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 201
                return response
            else:
                mensaje = f'{self.model.__name__} no se pudo registrar'
                error = form.errors
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 400
                return response
        else:
            return redirect('fichasNombres')

class editarFichaNombre(UpdateView):
    model = Nombreficha
    template_name = 'nombre_fichas/editar.html'
    form_class = NombreFichaForm
    success_url = reverse_lazy('fichasNombres')

    def post(self, request, *args, **kwargs):
        if is_ajax(request=request):
            form = self.form_class(request.POST, instance = self.get_object())
            if form.is_valid():
                form.save()
                mensaje = f'{self.model.__name__} actualizado correctamente!'
                error = 'no hay error'
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 201
                return response
            else:
                mensaje = f'{self.model.__name__} no se pudo actualizar'
                error = form.errors
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 400
                return response
        else:
            redirect('fichasNombres')

class confirmarEliminarFichaNombre(DeleteView):
    model = Nombreficha
    template_name = 'nombre_fichas/nombre_fichas_confirm_delete.html'
    success_url = reverse_lazy('fichasNombres')

    def post(self, request, *args, **kwargs):
        return render(request, self.template_name)

def eliminarFichaNombre(request, id):
    eliminar = Nombreficha.objects.get(id = id)
    eliminar.delete()
    return redirect('fichasNombres')
# ! Modulo de nombre de fichas


# ! Modulo de gallinas
class Gallinass(ListView):
    model = Gallinas
    template_name = 'gallinas/gallinas.html'

    def get_queryset(self):
        return self.model.objects.all().order_by('-id')

    def get_context_data(self, **kwargs):
        contexto = {}
        contexto["gallinas"] = self.get_queryset()
        return contexto

    def get(self, request, *args, **kwargs):
        gallinasMenu = True
        if is_ajax(request=request):
            return HttpResponse(serialize('json', self.get_context_data()), 'application/json')
        else:
            return render(request, self.template_name, {'gallinas': self.get_queryset(), 'gallinasMenu': gallinasMenu })

class crearGallinas(CreateView):
    model = Gallinas
    template_name = 'gallinas/crear.html'
    form_class = GallinasForm
    success_url = reverse_lazy('gallinas')

    def post(self, request, *args, **kwargs):
        if is_ajax(request=request):
            form = self.form_class(request.POST)
            if form.is_valid():
                galponForm = form.cleaned_data['id_galpon']
                galpon = Galpones.objects.get(id=galponForm.id)
                cant_gallinas_form = form.cleaned_data['cantidad_gallinas']
                galpon.cant_gall += cant_gallinas_form
                galpon.save()
                form.save()
                gallinasSaved = Gallinas.objects.latest('id')
                try:
                    registroDiarioSaved = Registrodiario.objects.filter(fecha=fechaActual(), id_galpon=galponForm).last()
                except ObjectDoesNotExist:
                    registroDiarioSaved = 0
                id_alimentacion = None
                id_producciondiaria = None
                id_mortades = None
                if hasattr(registroDiarioSaved, 'id_alimentacion'):
                    id_alimentacion = registroDiarioSaved.id_alimentacion
                if hasattr(registroDiarioSaved, 'id_producciondiaria'):
                    id_producciondiaria = registroDiarioSaved.id_producciondiaria
                if hasattr(registroDiarioSaved, 'id_mortades'):
                    id_mortades = registroDiarioSaved.id_mortades
                if id_alimentacion or id_producciondiaria:
                    registrosDiariosSavedAll = Registrodiario.objects.filter(id_galpon=galponForm, fecha=fechaActual())
                    for registro in registrosDiariosSavedAll:
                        registro.id_alimentacion = registroDiarioSaved.id_alimentacion
                        registro.id_gallinas = gallinasSaved
                        if hasattr(registroDiarioSaved, 'id_mortades'):
                            registro.id_mortades = registroDiarioSaved.id_mortades
                        registro.save()
                elif registroDiarioSaved:
                    if not hasattr(registroDiarioSaved, 'id_gallinas') or hasattr(registroDiarioSaved, 'id_gallinas'):
                        registroDiarioSaved.id_gallinas = gallinasSaved
                        registroDiarioSaved.save()
                else:
                    registro = Registrodiario(id_galpon=galponForm ,id_gallinas=gallinasSaved)
                    registro.save()
                mensaje = f'{self.model.__name__} registrado correctamente!'
                error = 'no hay error'
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 201
                return response
            else:
                mensaje = f'{self.model.__name__} no se pudo registrar'
                error = form.errors
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 400
                return response
        else:
            return redirect('gallinas')

class editarGallinas(UpdateView):
    model = Gallinas
    template_name = 'gallinas/editar.html'
    form_class = GallinasForm
    success_url = reverse_lazy('gallinas')

    def post(self, request, *args, **kwargs):
        if is_ajax(request=request):
            form = self.form_class(request.POST, instance = self.get_object())
            if form.is_valid():
                galponForm = form.cleaned_data['id_galpon'].id
                galpon = Galpones.objects.get(id=galponForm)
                cant_gallinas_form = form.cleaned_data['cantidad_gallinas']
                gallinasSaved = Gallinas.objects.get(id=self.get_object().id).cantidad_gallinas
                if gallinasSaved > cant_gallinas_form:
                    gallinasSaved -= cant_gallinas_form
                    galpon.cant_gall -= gallinasSaved
                else:
                    cant_gallinas_form -= gallinasSaved
                    galpon.cant_gall += cant_gallinas_form
                galpon.save()
                form.save()
                mensaje = f'{self.model.__name__} actualizado correctamente!'
                error = 'no hay error'
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 201
                return response
            else:
                mensaje = f'{self.model.__name__} no se pudo actualizar'
                error = form.errors
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 400
                return response
        else:
            redirect('gallinas')

class confirmarEliminarGallinas(DeleteView):
    model = Gallinas
    template_name = 'gallinas/gallinas_confirm_delete.html'
    success_url = reverse_lazy('gallinas')

    def post(self, request, *args, **kwargs):
        return render(request, self.template_name)

def eliminarGallinas(request, id):
    eliminar = Gallinas.objects.get(id = id)
    eliminar.delete()
    return redirect('gallinas')
# ! Modulo de gallinas


# ! Modulo de galpones
class Galponess(ListView):
    model = Galpones
    template_name = 'galpones/galpones.html'

    def get_queryset(self):
        select = self.request.GET.get('search')
        busqueda = self.request.GET.get('buscar')
        busquedaDate = self.request.GET.get('buscarDate')
        if select == 'date':
            query = self.model.objects.filter(
                Q(fecha__icontains = busquedaDate)
                ).distinct().order_by('-id')
            busqueda = ''
        elif select == 'input':
            query = self.model.objects.filter(
                Q(nombre_galpon__icontains = busqueda) |
                Q(area__icontains = busqueda) |
                Q(capac_bebed__icontains = busqueda) |
                Q(cant_bebed__icontains = busqueda) |
                Q(capac_comed__icontains = busqueda) |
                Q(cant_comed__icontains = busqueda) |
                Q(capac_gall__icontains = busqueda) |
                Q(cant_gall__icontains = busqueda) |
                Q(capac_nidales__icontains = busqueda) |
                Q(cant_nidales__icontains = busqueda)
            ).distinct()
            busquedaDate = ''
        else:
            query = self.model.objects.all().order_by('-id')
        return query

    def get_context_data(self, **kwargs):
        contexto = {}
        contexto["galpones"] = self.get_queryset()
        return contexto
    
    def get(self, request, *args, **kwargs):
        galponMenu = True
        if is_ajax(request=request):
                return HttpResponse(serialize('json', self.get_context_data()), 'application/json')
        else:
            return render(request, self.template_name, {'galpones': self.get_queryset(), 'galponMenu': galponMenu})

class crearGalpon(CreateView):
    model = Galpones
    template_name = 'galpones/crear.html'
    form_class = GalponesForm
    success_url = reverse_lazy('galpones')

    def post(self, request, *args, **kwargs):
        if is_ajax(request=request):
            form = self.form_class(request.POST)
            if form.is_valid():
                form.save()
                mensaje = f'{self.model.__name__} registrado correctamente!'
                error = 'no hay error'
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 201
                return response
            else:
                mensaje = f'{self.model.__name__} no se pudo registrar'
                error = form.errors
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 400
                return response
        else:
            return redirect('galpones')

class editarGalpon(UpdateView):
    model = Galpones
    template_name = 'galpones/editar.html'
    form_class = GalponesForm
    success_url = reverse_lazy('galpones')

    def post(self, request, *args, **kwargs):
        if is_ajax(request=request):
            form = self.form_class(request.POST, instance = self.get_object())
            if form.is_valid():
                form.save()
                mensaje = f'{self.model.__name__} actualizado correctamente!'
                error = 'no hay error'
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 201
                return response
            else:
                mensaje = f'{self.model.__name__} no se pudo actualizar'
                error = form.errors
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 400
                return response
        else:
            redirect('galpones')

class confirmarEliminarGalpon(DeleteView):
    model = Galpones
    template_name = 'galpones/galpones_confirm_delete.html'
    success_url = reverse_lazy('galpones')

    def post(self, request, *args, **kwargs):
        return render(request, self.template_name)

def eliminarGalpon(request, id):
    eliminar = Galpones.objects.get(id = id)
    eliminar.delete()
    return redirect('galpones')
# ! Modulo de galpones


# ! Modulo de jornadas
class Jornadass(ListView):
    model = Jornada
    template_name = 'jornadas/jornadas.html'

    def get_queryset(self):
        return self.model.objects.all().order_by('-id')

    def get_context_data(self, **kwargs):
        contexto = {}
        contexto["jornadas"] = self.get_queryset()
        return contexto
    
    def get(self, request, *args, **kwargs):
        user = Usuario.objects.filter(id = request.user.id).values_list('is_staff', flat = True)
        if user[0] == True:
            if is_ajax(request=request):
                return HttpResponse(serialize('json', self.get_context_data()), 'application/json')
            else:
                return render(request, self.template_name, {'jornadas': self.get_queryset()})
        else:
            return redirect('interfaz')

class crearJornada(CreateView):
    model = Jornada
    template_name = 'jornadas/crear.html'
    form_class = JornadaForm
    success_url = reverse_lazy('jornadas')

    def post(self, request, *args, **kwargs):
        if is_ajax(request=request):
            form = self.form_class(request.POST)
            if form.is_valid():
                form.save()
                mensaje = f'{self.model.__name__} registrado correctamente!'
                error = 'no hay error'
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 201
                return response
            else:
                mensaje = f'{self.model.__name__} no se pudo registrar'
                error = form.errors
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 400
                return response
        else:
            return redirect('jornadas')

class editarJornada(UpdateView):
    model = Jornada
    template_name = 'jornadas/editar.html'
    form_class = JornadaForm
    success_url = reverse_lazy('jornadas')

    def post(self, request, *args, **kwargs):
        if is_ajax(request=request):
            form = self.form_class(request.POST, instance = self.get_object())
            if form.is_valid():
                form.save()
                mensaje = f'{self.model.__name__} actualizado correctamente!'
                error = 'no hay error'
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 201
                return response
            else:
                mensaje = f'{self.model.__name__} no se pudo actualizar'
                error = form.errors
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 400
                return response
        else:
            redirect('jornadas')

class confirmarEliminarJornada(DeleteView):
    model = Jornada
    template_name = 'jornadas/jornadas_confirm_delete.html'
    success_url = reverse_lazy('jornadas')

    def post(self, request, *args, **kwargs):
        return render(request, self.template_name)

def eliminarJornada(request, id):
    eliminar = Jornada.objects.get(id = id)
    eliminar.delete()
    return redirect('jornadas')
# ! Modulo de jornadas


# ! Modulo de lineas
class Lineass(ListView):
    model = Linea
    template_name = 'lineas/lineas.html'

    def get_queryset(self):
        return self.model.objects.all().order_by('-id')

    def get_context_data(self, **kwargs):
        contexto = {}
        contexto["lineas"] = self.get_queryset()
        return contexto
    
    def get(self, request, *args, **kwargs):
        user = Usuario.objects.filter(id = request.user.id).values_list('is_staff', flat = True)
        if user[0] == True:
            if is_ajax(request=request):
                return HttpResponse(serialize('json', self.get_context_data()), 'application/json')
            else:
                return render(request, self.template_name, {'lineas': self.get_queryset()})
        else:
            return redirect('interfaz')

class crearLinea(CreateView):
    model = Linea
    template_name = 'lineas/crear.html'
    form_class = LineaForm
    success_url = reverse_lazy('lineas')

    def post(self, request, *args, **kwargs):
        if is_ajax(request=request):
            form = self.form_class(request.POST)
            if form.is_valid():
                form.save()
                mensaje = f'{self.model.__name__} registrado correctamente!'
                error = 'no hay error'
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 201
                return response
            else:
                mensaje = f'{self.model.__name__} no se pudo registrar'
                error = form.errors
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 400
                return response
        else:
            return redirect('lineas')

class editarLinea(UpdateView):
    model = Linea
    template_name = 'lineas/editar.html'
    form_class = LineaForm
    success_url = reverse_lazy('lineas')

    def post(self, request, *args, **kwargs):
        if is_ajax(request=request):
            form = self.form_class(request.POST, instance = self.get_object())
            if form.is_valid():
                form.save()
                mensaje = f'{self.model.__name__} actualizado correctamente!'
                error = 'no hay error'
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 201
                return response
            else:
                mensaje = f'{self.model.__name__} no se pudo actualizar'
                error = form.errors
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 400
                return response
        else:
            redirect('lineas')

class confirmarEliminarLinea(DeleteView):
    model = Linea
    template_name = 'lineas/lineas_confirm_delete.html'
    success_url = reverse_lazy('lineas')

    def post(self, request, *args, **kwargs):
        return render(request, self.template_name)

def eliminarLinea(request, id):
    eliminar = Linea.objects.get(id = id)
    eliminar.delete()
    return redirect('lineas')
# ! Modulo de lineas


# ! Modulo de mortalidad y descarte
class Mortalidadd(ListView):
    model = MortalidadDescarte
    template_name = 'mortalidad_descarte/mortalidad_descarte.html'

    def get_queryset(self):
        select = self.request.GET.get('search')
        busqueda = self.request.GET.get('buscar')
        busquedaDate = self.request.GET.get('buscarDate')
        if select == 'date':
            query = self.model.objects.filter(
                Q(fecha__icontains = busquedaDate)
                ).distinct().order_by('-id')
            busqueda = ''
        elif select == 'input':
            query = self.model.objects.filter(
                Q(id_galpon__nombre_galpon__icontains = busqueda) |
                Q(cant_muertas__icontains = busqueda) |
                Q(cant_descarte__icontains = busqueda) |
                Q(id_tipo_descarte__tipo__icontains = busqueda) |
                Q(saldo__icontains = busqueda)
            ).distinct()
            busquedaDate = ''
        else:
            query = self.model.objects.all().order_by('-id')
        return query

    def get_context_data(self, **kwargs):
        contexto = {}
        contexto["mortalidad_descarte"] = self.get_queryset()
        contexto["data"] = Galpones.objects.all().values_list("id", "cant_gall")
        return contexto
    
    def get(self, request, *args, **kwargs):
        mortaDesMenu = True
        if is_ajax(request=request):
                return HttpResponse(serialize('json', self.get_context_data()), 'application/json')
        else:
            return render(request, self.template_name, {'mortalidad_descarte': self.get_queryset(), 'mortaDesMenu': mortaDesMenu})

class crearMortalidad(CreateView):
    model = MortalidadDescarte
    template_name = 'mortalidad_descarte/crear.html'
    form_class = MortalidadDescarteForm
    success_url = reverse_lazy('mortalidad_descarte')

    def post(self, request, *args, **kwargs):
        if is_ajax(request=request):
            form = self.form_class(request.POST)
            if form.is_valid():
                cantGallinas = form.cleaned_data['saldo']
                galponForm = form.cleaned_data['id_galpon']
                galponSaved = Galpones.objects.get(id=galponForm.id)
                galponSaved.cant_gall = cantGallinas
                galponSaved.save()
                form.save()
                mortaDesSaved = MortalidadDescarte.objects.latest('id')
                try:
                    registroDiarioSaved = Registrodiario.objects.filter(fecha=fechaActual(), id_galpon=galponForm).last()
                except ObjectDoesNotExist:
                    registroDiarioSaved = 0
                id_alimentacion = None
                id_producciondiaria = None
                id_gallinas = None
                if hasattr(registroDiarioSaved, 'id_alimentacion'):
                    id_alimentacion = registroDiarioSaved.id_alimentacion
                if hasattr(registroDiarioSaved, 'id_producciondiaria'):
                    id_producciondiaria = registroDiarioSaved.id_producciondiaria
                if hasattr(registroDiarioSaved, 'id_gallinas'):
                    id_gallinas = registroDiarioSaved.id_gallinas
                if id_alimentacion and id_producciondiaria:
                    registrosDiariosSavedAll = Registrodiario.objects.filter(id_galpon=galponForm, fecha=fechaActual())
                    for registro in registrosDiariosSavedAll:
                        registro.id_gallinas = registroDiarioSaved.id_gallinas
                        registro.id_alimentacion = registroDiarioSaved.id_alimentacion
                        registro.id_mortades = mortaDesSaved
                        registro.save()
                elif registroDiarioSaved:
                    if not hasattr(registroDiarioSaved, 'id_mortades') or hasattr(registroDiarioSaved, 'id_mortades'):
                        registroDiarioSaved.id_mortades = mortaDesSaved
                        registroDiarioSaved.save()
                else:
                    registro = Registrodiario(id_galpon=galponForm ,id_mortades=mortaDesSaved)
                    registro.save()
                mensaje = f'{self.model.__name__} registrado correctamente!'
                error = 'no hay error'
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 201
                return response
            else:
                mensaje = f'{self.model.__name__} no se pudo registrar'
                error = form.errors
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 400
                return response
        else:
            return redirect('mortalidad_descarte')

class editarMortalidad(UpdateView):
    model = MortalidadDescarte
    template_name = 'mortalidad_descarte/editar.html'
    form_class = MortalidadDescarteForm
    success_url = reverse_lazy('mortalidad_descarte')

    def post(self, request, *args, **kwargs):
        if is_ajax(request=request):
            form = self.form_class(request.POST, instance = self.get_object())
            if form.is_valid():
                form.save()
                mensaje = f'{self.model.__name__} actualizado correctamente!'
                error = 'no hay error'
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 201
                return response
            else:
                mensaje = f'{self.model.__name__} no se pudo actualizar'
                error = form.errors
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 400
                return response
        else:
            redirect('mortalidad_descarte')

class confirmarEliminarMortalidad(DeleteView):
    model = MortalidadDescarte
    template_name = 'mortalidad_descarte/mortalidad_confirm_delete.html'
    success_url = reverse_lazy('mortalidad_descarte')

    def post(self, request, *args, **kwargs):
        return render(request, self.template_name)

def eliminarMortalidad(request, id):
    eliminar = MortalidadDescarte.objects.get(id = id)
    eliminar.delete()
    return redirect('mortalidad_descarte')
# ! Modulo de mortalidad y descarte


# ! Modulo de produccion diaria
class ProduccionDiariaa(ListView):
    model = ProduccionDiaria
    template_name = 'prod_diaria/prod_diaria.html'

    def get_queryset(self):
        select = self.request.GET.get('search')
        busqueda = self.request.GET.get('buscar')
        busquedaDate = self.request.GET.get('buscarDate')
        if select == 'date':
            query = self.model.objects.filter(
                Q(fecha__icontains = busquedaDate)
                ).distinct().order_by('-id')
            busqueda = ''
        elif select == 'input':
            query = self.model.objects.filter(
                Q(id_galpon__nombre_galpon__icontains = busqueda) |
                Q(id_jornada__jornada__icontains = busqueda) |
                Q(id_tipo_huevo__tipos_huevos = busqueda )|
                Q(id_usuario__nombre__icontains = busqueda)
            ).distinct().order_by('-id')
            busquedaDate = ''
        else:
            query = self.model.objects.filter(fecha=fechaActual()).order_by('-id')
        return query

    def get_context_data(self, **kwargs):
        contexto = {}
        contexto["produccion_diaria"] = self.get_queryset()
        return contexto
    
    def get(self, request, *args, **kwargs):
        prodDiariaMenu = True
        dataProd = ProduccionDiaria.objects.filter(fecha=fechaActual()).values()
        totalHuevos = 0
        totalRotos = 0
        totalDescarte = 0
        for dato in dataProd:
            totalHuevos += int(dato['cantidad'])
            totalRotos += int(dato['rotos'])
            totalDescarte += int(dato['descarte'])

        # ? Por dia        
        # totalTiposHuevos = TiposHuevos.objects.all()
        # for tipoHuevo in totalTiposHuevos:
        #     tipoHuevo.cantidad = 0
        #     tipoHuevo.percent = 0
        #     for dato in self.get_queryset():
        #         if dato.id_tipo_huevo == tipoHuevo:
        #             tipoHuevo.cantidad += dato.cantidad
        #     tipoHuevo.porc = round((tipoHuevo.cantidad * 100) / totalHuevos, 2)
        
        # ? Por semana
        prodSemana = []
        d = datetime.datetime.now()
        d -= timedelta(days=d.weekday())
        totalHSemana = 0
        for x in range(0, 7):
            prodDiariaDate = ProduccionDiaria.objects.filter(fecha=d)
            if prodDiariaDate:
                for dato in prodDiariaDate:
                    prodSemana.append(dato)
                    totalHSemana += dato.cantidad
            d += timedelta(days=1)

        totalTiposHuevos = TiposHuevos.objects.all()
        for tipoHuevo in totalTiposHuevos:
            tipoHuevo.cantidad = 0
            tipoHuevo.percent = 0
            for dato in prodSemana:
                if dato.id_tipo_huevo == tipoHuevo:
                    tipoHuevo.cantidad += dato.cantidad
            tipoHuevo.porc = round((tipoHuevo.cantidad * 100) / totalHSemana, 2)
            d += timedelta(days=1)
        
        if is_ajax(request=request):
            return HttpResponse(serialize('json', self.get_context_data()), 'application/json')
        else:
            return render(request, self.template_name, {'produccion_diaria': self.get_queryset(), 'prodDiariaMenu': prodDiariaMenu, 'totalHuevos': totalHuevos, 'totalRotos': totalRotos, 'totalDescarte': totalDescarte, 'totalTiposHuevos': totalTiposHuevos})
    
    def post(self, request, *args, **kwargs):
        prodDiariaMenu = True
        query = self.get_queryset()
        if not query:
            messages.error(request, 'Debes buscar algun dato para generar el reporte')
            return render(request, self.template_name, {'produccion_diaria': self.get_queryset(), 'prodDiariaMenu': prodDiariaMenu})

        wb = Workbook()
        ws = wb.active
        #nombre de la hoja de excel
        ws.title = f'Reporte de {self.model.nameTitle()}'

        #configutación del encabezado
        ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
        ws['B2'].border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['B2'].fill = PatternFill(start_color='39A900', fill_type='solid')
        ws['B2'].font = Font(name='Arial', size=15, bold=True, color='FFFFFF')
        ws['B2'] = f'REPORTE {self.model._meta.verbose_name.upper()}'
        
        ws.merge_cells('B2:I2')
        listColumn = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
        listName = ['Galpón', 'Jornada', 'Tipo de Huevo', 'Cantidad', 'Rotos', 'Descarte', 'Usuario', 'Fecha']
        countName = 0
        count = 3
        ws.row_dimensions[2].height = 25

        
        for i in listColumn:
            ws.column_dimensions[i].width = 35
            ws[f'{listColumn[countName]}3'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'{listColumn[countName]}3'].border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                                                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            ws[f'{listColumn[countName]}3'].fill = PatternFill(start_color='FFCE40', fill_type='solid')
            ws[f'{listColumn[countName]}3'].font = Font(name='Arial', size=11)
            ws[f'{listColumn[countName]}3'] = listName[countName]
            count += 1
            countName += 1
            # Pintamos los datos en el reporte
        listName = ['id_galpon', 'id_jornada', 'id_tipo_huevo', 'cantidad', 'rotos', 'descarte', 'id_usuario', 'fecha']
        countColumn = 2
        for i in listName:
            countRow = 4
            for q in query:
                ws.cell(row=countRow, column=countColumn).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=countRow, column=countColumn).border = Border(left=Side(border_style='thin'),
                                                                        right=Side(border_style='thin'),
                                                                        top=Side(border_style='thin'),
                                                                        bottom=Side(border_style='thin'))
                ws.cell(row=countRow, column=countColumn).fill = PatternFill(start_color='FBFBE2', fill_type='solid')
                ws.cell(row=countRow, column=countColumn).font = Font(name='Arial', size='11')
                # Obtener el valor de la columna

                # Obtener el valor de la columna
                if i == 'id_usuario':
                    valueRow = getattr(q.id_usuario, 'nombre', 'nombre') if q.id_usuario else ''
                elif i == 'id_galpon':
                    valueRow = getattr(q.id_galpon, 'nombre_galpon', 'nombre_galpon') if q.id_galpon else ''
                elif i == 'id_jornada':
                    valueRow = getattr(q.id_jornada, 'jornada', 'jornada') if q.id_jornada else ''
                elif i == 'id_tipo_huevo':
                    valueRow = getattr(q.id_tipo_huevo, 'tipos_huevos', 'tipos_huevos') if q.id_tipo_huevo else ''
                elif i == 'fecha':
                    valueRow = getattr(q, i).strftime('%Y-%m-%d') if getattr(q, i) else ''
                else:
                    valueRow = getattr(q, i)

                
                ws.cell(row=countRow, column=countColumn).value = valueRow
                countRow += 1
            countColumn += 1
        
        dataProd = ProduccionDiaria.objects.filter(fecha=fechaActual()).values()
        totalHuevos = 0
        totalRotos = 0
        totalDescarte = 0
        for dato in dataProd:
            totalHuevos += int(dato['cantidad'])
            totalRotos += int(dato['rotos'])
            totalDescarte += int(dato['descarte'])
        
        ws[f'B{countRow}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'B{countRow}'].border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                                                        top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws[f'B{countRow}'].fill = PatternFill(start_color='FFCE40', fill_type='solid')
        ws[f'B{countRow}'].font = Font(name='Arial', size=11)
        ws.cell(row=countRow, column=2).value = 'Total de huevos buenos, rotos y descarte en el día'
        ws.merge_cells(f'B{countRow}:D{countRow}')

        ws[f'E{countRow}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'E{countRow}'].border = Border(left=Side(border_style='thin'),
                                                                right=Side(border_style='thin'),
                                                                top=Side(border_style='thin'),
                                                                bottom=Side(border_style='thin'))
        ws[f'E{countRow}'].fill = PatternFill(start_color='FBFBE2', fill_type='solid')
        ws[f'E{countRow}'].font = Font(name='Arial', size='11')
        ws.cell(row=countRow, column=5).value = totalHuevos

        ws[f'F{countRow}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'F{countRow}'].border = Border(left=Side(border_style='thin'),
                                                                right=Side(border_style='thin'),
                                                                top=Side(border_style='thin'),
                                                                bottom=Side(border_style='thin'))
        ws[f'F{countRow}'].fill = PatternFill(start_color='FBFBE2', fill_type='solid')
        ws[f'F{countRow}'].font = Font(name='Arial', size='11')
        ws.cell(row=countRow, column=6).value = totalRotos

        ws[f'G{countRow}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'G{countRow}'].border = Border(left=Side(border_style='thin'),
                                                                right=Side(border_style='thin'),
                                                                top=Side(border_style='thin'),
                                                                bottom=Side(border_style='thin'))
        ws[f'G{countRow}'].fill = PatternFill(start_color='FBFBE2', fill_type='solid')
        ws[f'G{countRow}'].font = Font(name='Arial', size='11')
        ws.cell(row=countRow, column=7).value = totalDescarte

        # ? Por semana
        prodSemana = []
        d = datetime.datetime.now()
        d -= timedelta(days=d.weekday())
        totalHSemana = 0
        for x in range(0, 7):
            prodDiariaDate = ProduccionDiaria.objects.filter(fecha=d)
            if prodDiariaDate:
                for dato in prodDiariaDate:
                    prodSemana.append(dato)
                    totalHSemana += dato.cantidad
            d += timedelta(days=1)

        totalTiposHuevos = TiposHuevos.objects.all()
        for tipoHuevo in totalTiposHuevos:
            tipoHuevo.cantidad = 0
            tipoHuevo.percent = 0
            for dato in prodSemana:
                if dato.id_tipo_huevo == tipoHuevo:
                    tipoHuevo.cantidad += dato.cantidad
            tipoHuevo.porc = round((tipoHuevo.cantidad * 100) / totalHSemana, 2)
            d += timedelta(days=1)

        listTitle = ['Tipo de Huevo', 'Total de Huevos', 'Porcentaje semanal']
        listColumn = ['B', 'D', 'G']
        counter = 0
        for i in listTitle:
            countRowValue = countRow
            ws[f'{listColumn[counter]}{countRow + 2}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'{listColumn[counter]}{countRow + 2}'].border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                                                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            ws[f'{listColumn[counter]}{countRow + 2}'].fill = PatternFill(start_color='FFCE40', fill_type='solid')
            ws[f'{listColumn[counter]}{countRow + 2}'].font = Font(name='Arial', size=11)

            if i == 'Tipo de Huevo':
                ws.cell(row=countRow + 2, column=2).value = i
                ws.merge_cells(f'B{countRow + 2}:C{countRow + 2}')
            elif i == 'Total de Huevos':
                ws.cell(row=countRow + 2, column=4).value = i
                ws.merge_cells(f'D{countRow + 2}:F{countRow + 2}')
            else:
                ws.cell(row=countRow + 2, column=7).value = i
                ws.merge_cells(f'G{countRow + 2}:I{countRow + 2}')

            for j in totalTiposHuevos:
                ws[f'{listColumn[counter]}{countRowValue + 3}'].alignment = Alignment(horizontal='center', vertical='center')
                ws[f'{listColumn[counter]}{countRowValue + 3}'].border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                                                    top=Side(border_style='thin'),
                                                    bottom=Side(border_style='thin'))
                ws[f'{listColumn[counter]}{countRowValue + 3}'].fill = PatternFill(start_color='FBFBE2', fill_type='solid')
                ws[f'{listColumn[counter]}{countRowValue + 3}'].font = Font(name='Arial', size='11')
                if i == 'Tipo de Huevo':
                    ws.cell(row=countRowValue + 3, column=2).value = j.tipos_huevos
                    ws.merge_cells(f'B{countRowValue + 3}:C{countRowValue + 3}')
                elif i == 'Total de Huevos':
                    ws.cell(row=countRowValue + 3, column=4).value = j.cantidad
                    ws.merge_cells(f'D{countRowValue + 3}:F{countRowValue + 3}')
                else:
                    ws.cell(row=countRowValue + 3, column=7).value = f'{j.porc}%'
                    ws.merge_cells(f'G{countRowValue + 3}:I{countRowValue + 3}')
                countRowValue += 1
            counter += 1

        # Nombre del archivo
        nombreArchivo = f'REPORTE {self.model.nameTitle().upper()}.xlsx'
        # Definir el tipo de respuesta
        response = HttpResponse(content_type='application/ms-excel')
        contenido = "attachment; filename={0}".format(nombreArchivo)
        response['Content-Disposition'] = contenido
        wb.save(response)
        return response

class crearProdDiaria(LoginRequiredMixin, CreateView):
    model = ProduccionDiaria
    template_name = 'prod_diaria/crear.html'
    form_class = ProduccionDiariaForm
    success_url = reverse_lazy('produccion_diaria')

    # se crea el metodo From_valid para poder asignar automaticamente el usuario actual al campo 'id_usuario'
    # modificacion -> y se excluye el campo 'id_usurio' en forms.py 
    def form_valid(self, form):
        form.instance.id_usuario = self.request.user  # Asignar el usuario actual al campo id_usuario
        return super().form_valid(form)
    
    def post(self, request, *args, **kwargs):
        if is_ajax(request=request):
            form = self.form_class(request.POST)
            if form.is_valid():
                form.instance.id_usuario = self.request.user
                galponForm = form.cleaned_data['id_galpon']
                form.save()
                prodDiariaSaved = ProduccionDiaria.objects.latest('id')
                try:
                    registroDiarioSaved = Registrodiario.objects.filter(fecha=fechaActual(), id_galpon=galponForm).last()
                except ObjectDoesNotExist:
                    registroDiarioSaved = 0
                id_alimentacion = None
                id_mortades = None
                id_gallinas = None
                if hasattr(registroDiarioSaved, 'id_alimentacion'):
                    id_alimentacion = registroDiarioSaved.id_alimentacion
                if hasattr(registroDiarioSaved, 'id_mortades'):
                    id_mortades = registroDiarioSaved.id_mortades
                if hasattr(registroDiarioSaved, 'id_gallinas'):
                    id_gallinas = registroDiarioSaved.id_gallinas
                if id_alimentacion or id_gallinas:
                    registrosDiariosSavedAll = Registrodiario.objects.filter(id_galpon=galponForm, fecha=fechaActual())
                    registro = Registrodiario.objects.create(
                        id_galpon=registroDiarioSaved.id_galpon,
                        id_gallinas=id_gallinas,
                        id_producciondiaria=prodDiariaSaved,
                        id_mortades=id_mortades,
                        id_alimentacion=id_alimentacion
                    )
                    registro.save()
                    for registro in registrosDiariosSavedAll:
                        registro.id_alimentacion = registroDiarioSaved.id_alimentacion
                        if hasattr(registroDiarioSaved, 'id_gallinas'):
                            registro.id_gallinas = id_gallinas
                        if hasattr(registroDiarioSaved, 'id_mortades'):
                            registro.id_mortades = registroDiarioSaved.id_mortades
                        registro.save()
                elif registroDiarioSaved:
                    if not hasattr(registroDiarioSaved, 'id_producciondiaria'):
                        registroDiarioSaved.id_producciondiaria = prodDiariaSaved
                        registroDiarioSaved.save()
                    else:
                        registro = Registrodiario.objects.create(
                            id_galpon=registroDiarioSaved.id_galpon,
                            id_gallinas=id_gallinas,
                            id_producciondiaria=prodDiariaSaved,
                            id_mortades=id_mortades,
                            id_alimentacion=id_alimentacion
                        )
                        registro.save()
                else:
                    registro = Registrodiario(id_galpon=galponForm ,id_producciondiaria=prodDiariaSaved)
                    registro.save()
                try:
                    alimentacionSaved = Alimentacion.objects.filter(id_galpon=galponForm, fecha=fechaActual()).last()
                except ObjectDoesNotExist:
                    alimentacionSaved = 0
                if alimentacionSaved:
                    dataProd = ProduccionDiaria.objects.filter(id_galpon=galponForm, fecha=fechaActual()).values()
                    totalHuevos = 0
                    for dato in dataProd:
                        totalHuevos += int(dato['cantidad'])
                    alimentacionSaved.c_a = float(alimentacionSaved.kg_total) / (totalHuevos / 12)
                    alimentacionSaved.save()
                mensaje = f'{self.model.__name__} registrado correctamente!'
                error = 'no hay error'
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 201
                return response
            else:
                mensaje = f'{self.model.__name__} no se pudo registrar'
                error = form.errors
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 400
                return response
        else:
            return redirect('produccion_diaria')

class editarProdDiaria(UpdateView):
    model = ProduccionDiaria
    template_name = 'prod_diaria/editar.html'
    form_class = ProduccionDiariaForm
    success_url = reverse_lazy('produccion_diaria')

    def post(self, request, *args, **kwargs):
        if is_ajax(request=request):
            form = self.form_class(request.POST, instance = self.get_object())
            if form.is_valid():
                galponForm = form.cleaned_data.get('id_galpon').id
                try:
                    alimentacionSaved = Alimentacion.objects.get(id_galpon=galponForm, fecha=fechaActual())
                except ObjectDoesNotExist:
                    alimentacionSaved = 0
                if alimentacionSaved:
                    dataProdSaved = ProduccionDiaria.objects.filter(id_galpon=galponForm, fecha=fechaActual()).values()
                    cantidadProdSaved = ProduccionDiaria.objects.get(id=self.get_object().id).cantidad
                    cantidadProdForm = form.cleaned_data.get('cantidad')
                    totalHuevos = 0
                    for dato in dataProdSaved:
                        totalHuevos += int(dato['cantidad'])
                    if cantidadProdSaved > cantidadProdForm:
                        cantidadProdSaved -= cantidadProdForm
                        totalHuevos -= cantidadProdSaved
                    else:
                        cantidadProdForm -= cantidadProdSaved
                        totalHuevos += cantidadProdForm
                    alimentacionSaved.c_a = float(alimentacionSaved.kg_total) / (totalHuevos / 12)
                    alimentacionSaved.save()
                form.save()
                mensaje = f'{self.model.__name__} actualizado correctamente!'
                error = 'no hay error'
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 201
                return response
            else:
                mensaje = f'{self.model.__name__} no se pudo actualizar'
                error = form.errors
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 400
                return response
        else:
            return redirect('produccion_diaria')

class confirmarEliminarProdDiaria(DeleteView):
    model = ProduccionDiaria
    template_name = 'prod_diaria/prod_diaria_confirm_delete.html'
    success_url = reverse_lazy('produccion_diaria')

    def post(self, request, *args, **kwargs):
        return render(request, self.template_name)

def eliminarProdDiaria(request, id):
    eliminar = ProduccionDiaria.objects.get(id = id)
    eliminar.delete()
    return redirect('produccion_diaria')
# ! Modulo de produccion diaria


# ! Modulo de rol
class Roll(ListView):
    model = Rol
    template_name = 'rol/rol.html'

    def get_queryset(self):
        return self.model.objects.all().order_by('-id')

    def get_context_data(self, **kwargs):
        contexto = {}
        contexto["rol"] = self.get_queryset()
        return contexto
    
    def get(self, request, *args, **kwargs):
        user = Usuario.objects.filter(id = request.user.id).values_list('is_staff', flat = True)
        if user[0] == True:
            if is_ajax(request=request):
                return HttpResponse(serialize('json', self.get_context_data()), 'application/json')
            else:
                return render(request, self.template_name, {'rol': self.get_queryset()})
        else:
            return redirect('interfaz')

class crearRol(CreateView):
    model = Rol
    template_name = 'rol/crear.html'
    form_class = RolForm
    success_url = reverse_lazy('rol')

    def post(self, request, *args, **kwargs):
        if is_ajax(request=request):
            form = self.form_class(request.POST)
            if form.is_valid():
                form.save()
                mensaje = f'{self.model.__name__} registrado correctamente!'
                error = 'no hay error'
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 201
                return response
            else:
                mensaje = f'{self.model.__name__} no se pudo registrar'
                error = form.errors
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 400
                return response
        else:
            return redirect('rol')

class editarRol(UpdateView):
    model = Rol
    template_name = 'rol/editar.html'
    form_class = RolForm
    success_url = reverse_lazy('rol')

    def post(self, request, *args, **kwargs):
        if is_ajax(request=request):
            form = self.form_class(request.POST, instance = self.get_object())
            if form.is_valid():
                form.save()
                mensaje = f'{self.model.__name__} actualizado correctamente!'
                error = 'no hay error'
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 201
                return response
            else:
                mensaje = f'{self.model.__name__} no se pudo actualizar'
                error = form.errors
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 400
                return response
        else:
            redirect('rol')

class confirmarEliminarRol(DeleteView):
    model = Rol
    template_name = 'rol/rol_confirm_delete.html'
    success_url = reverse_lazy('rol')

    def post(self, request, *args, **kwargs):
        return render(request, self.template_name)

def eliminarRol(request, id):
    eliminar = Rol.objects.get(id = id)
    eliminar.delete()
    return redirect('rol')
# ! Modulo de rol


# ! Modulo de tipo de documento
class TipoDocc(ListView):
    model = TipoDoc
    template_name = 'tipo_doc/tipo_doc.html'

    def get_queryset(self):
        return self.model.objects.all().order_by('-id')

    def get_context_data(self, **kwargs):
        contexto = {}
        contexto["tipo_doc"] = self.get_queryset()
        return contexto
    
    def get(self, request, *args, **kwargs):
        user = Usuario.objects.filter(id = request.user.id).values_list('is_staff', flat = True)
        if user[0] == True:
            if is_ajax(request=request):
                return HttpResponse(serialize('json', self.get_context_data()), 'application/json')
            else:
                return render(request, self.template_name, {'tipo_doc': self.get_queryset()})
        else:
            return redirect('interfaz')

class crearTipoDoc(CreateView):
    model = TipoDoc
    template_name = 'tipo_doc/crear.html'
    form_class = TipoDocForm
    success_url = reverse_lazy('tipo_doc')

    def post(self, request, *args, **kwargs):
        if is_ajax(request=request):
            form = self.form_class(request.POST)
            if form.is_valid():
                form.save()
                mensaje = f'{self.model.__name__} registrado correctamente!'
                error = 'no hay error'
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 201
                return response
            else:
                mensaje = f'{self.model.__name__} no se pudo registrar'
                error = form.errors
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 400
                return response
        else:
            return redirect('tipo_doc')

class editarTipoDoc(UpdateView):
    model = TipoDoc
    template_name = 'tipo_doc/editar.html'
    form_class = TipoDocForm
    success_url = reverse_lazy('tipo_doc')

    def post(self, request, *args, **kwargs):
        if is_ajax(request=request):
            form = self.form_class(request.POST, instance = self.get_object())
            if form.is_valid():
                form.save()
                mensaje = f'{self.model.__name__} actualizado correctamente!'
                error = 'no hay error'
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 201
                return response
            else:
                mensaje = f'{self.model.__name__} no se pudo actualizar'
                error = form.errors
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 400
                return response
        else:
            redirect('tipo_doc')

class confirmarEliminarTipoDoc(DeleteView):
    model = TipoDoc
    template_name = 'tipo_doc/tipo_doc_confirm_delete.html'
    success_url = reverse_lazy('tipo_doc')

    def post(self, request, *args, **kwargs):
        return render(request, self.template_name)

def eliminarTipoDoc(request, id):
    eliminar = TipoDoc.objects.get(id = id)
    eliminar.delete()
    return redirect('tipo_doc')
# ! Modulo de tipo de documento


# ! Modulo de tipos de huevos
class TiposHuevoss(ListView):
    model = TiposHuevos
    template_name = 'tipos_huevos/tipos_huevos.html'

    def get_queryset(self):
        return self.model.objects.all().order_by('-id')

    def get_context_data(self, **kwargs):
        contexto = {}
        contexto["tipos_huevos"] = self.get_queryset()
        return contexto
    
    def get(self, request, *args, **kwargs):
        user = Usuario.objects.filter(id = request.user.id).values_list('is_staff', flat = True)
        if user[0] == True:
            if is_ajax(request=request):
                return HttpResponse(serialize('json', self.get_context_data()), 'application/json')
            else:
                return render(request, self.template_name, {'tipos_huevos': self.get_queryset()})
        else:
            return redirect('interfaz')

class crearTipoHuevo(CreateView):
    model = TiposHuevos
    template_name = 'tipos_huevos/crear.html'
    form_class = TiposHuevosForm
    success_url = reverse_lazy('tipos_huevos')

    def post(self, request, *args, **kwargs):
        if is_ajax(request=request):
            form = self.form_class(request.POST)
            if form.is_valid():
                form.save()
                mensaje = f'{self.model.__name__} registrado correctamente!'
                error = 'no hay error'
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 201
                return response
            else:
                mensaje = f'{self.model.__name__} no se pudo registrar'
                error = form.errors
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 400
                return response
        else:
            return redirect('tipos_huevos')

class editarTipoHuevo(UpdateView):
    model = TiposHuevos
    template_name = 'tipos_huevos/editar.html'
    form_class = TiposHuevosForm
    success_url = reverse_lazy('tipos_huevos')

    def post(self, request, *args, **kwargs):
        if is_ajax(request=request):
            form = self.form_class(request.POST, instance = self.get_object())
            if form.is_valid():
                form.save()
                mensaje = f'{self.model.__name__} actualizado correctamente!'
                error = 'no hay error'
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 201
                return response
            else:
                mensaje = f'{self.model.__name__} no se pudo actualizar'
                error = form.errors
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 400
                return response
        else:
            redirect('tipos_huevos')

class confirmarEliminarTipoHuevo(DeleteView):
    model = TiposHuevos
    template_name = 'tipos_huevos/tipos_huevos_confirm_delete.html'
    success_url = reverse_lazy('tipos_huevos')

    def post(self, request, *args, **kwargs):
        return render(request, self.template_name)

def eliminarTipoHuevo(request, id):
    eliminar = TiposHuevos.objects.get(id = id)
    eliminar.delete()
    return redirect('tipos_huevos')
# ! Modulo de tipos de huevos


# ! Modulo de usuario
class Usuarioss(ListView):
    model = Usuario
    template_name = 'usuarios/usuarios.html'

    def get_queryset(self):
        select = self.request.GET.get('search')
        busqueda = self.request.GET.get('buscar')
        busquedaDate = self.request.GET.get('buscarDate')
        if select == 'date':
            query = self.model.objects.filter(
                Q(registro__icontains = busquedaDate)
                ).distinct().order_by('-id')
            busqueda = ''
        elif select == 'input':
            query = self.model.objects.filter(
                Q(nombre__icontains = busqueda) |
                Q(apellido__icontains = busqueda) |
                Q(id_tipo_doc__tipo_doc__icontains = busqueda) |
                Q(documento__icontains = busqueda) |
                Q(celular__icontains = busqueda) |
                Q(id_ficha__num_ficha__icontains = busqueda) |
                Q(id_rol__tipo_rol__icontains = busqueda) |
                Q(email__icontains = busqueda)
            ).distinct().order_by('-id')
            busquedaDate = ''
        else:
            query = 0
        return query

    def get_context_data(self, **kwargs):
        contexto = {}
        contexto["usuarios"] = self.get_queryset()
        return contexto

    def get(self, request, *args, **kwargs):
        searchUser = True
        user = Usuario.objects.filter(id = request.user.id).values_list('is_staff', flat = True)
        if user[0] == True:
            if is_ajax(request=request):
                return HttpResponse(serialize('json', self.get_queryset()), 'application/json')
            else:
                return render(request, self.template_name, {'usuarios': self.get_queryset(), 'searchUser': searchUser})
        else:
            return redirect('interfaz')

    def post(self, request, *args, **kwargs):
        query = self.get_queryset()
        if not query:
            messages.error(request, 'Debes buscar algun dato para generar el reporte')
            return render(request, self.template_name, {'usuarios': self.get_queryset()})
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Hoja1'

            ws['B2'].alignment = Alignment(horizontal = 'center', vertical = 'center')
            ws['B2'].border = Border(left = Side(border_style = 'thin'), right = Side(border_style = 'thin'),
                                        top = Side(border_style = 'thin'), bottom = Side(border_style = 'thin'))
            ws['B2'].fill = PatternFill(start_color = '39A900', fill_type = 'solid')
            ws['B2'].font = Font(name = 'Arial', size = 15, bold = True, color = 'FFFFFF')
            ws['B2'] = f'REPORTE {self.model.__name__.upper()}'

            ws.merge_cells('B2:K2')
            listColumn = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
            listName = ['Nombre', 'Apellido', 'Tipo de Documento', 'Documento', 'Celular', 'email', 'Ficha', 'Rol', 'Registro', 'Ultima Conexión']
            countName = 0
            count = 3
            ws.row_dimensions[2].height = 25
            for i in listColumn:
                ws.column_dimensions[i].width = 35
                ws[f'{listColumn[countName]}3'].alignment = Alignment(horizontal = 'center', vertical = 'center')
                ws[f'{listColumn[countName]}3'].border = Border(left = Side(border_style = 'thin'), right = Side(border_style = 'thin'),
                                            top = Side(border_style = 'thin'), bottom = Side(border_style = 'thin'))
                ws[f'{listColumn[countName]}3'].fill = PatternFill(start_color = 'FFCE40', fill_type = 'solid')
                ws[f'{listColumn[countName]}3'].font = Font(name = 'Arial', size = 11)
                ws[f'{listColumn[countName]}3'] = listName[countName]
                count += 1
                countName += 1
            
            # Pintamos los datos en el reporte
            listName = ['nombre', 'apellido', 'id_tipo_doc', 'documento' , 'celular', 'email', 'id_ficha', 'id_rol', 'registro', 'last_login']
            countColumn = 2
            for i in listName:
                countRow = 4
                for q in query:
                    ws.cell(row = countRow, column = countColumn).alignment = Alignment(horizontal = 'center', vertical = 'center')
                    ws.cell(row = countRow, column = countColumn).border = Border(left = Side(border_style = 'thin'), right = Side(border_style = 'thin'),
                                                    top = Side(border_style = 'thin'), bottom = Side(border_style = 'thin'))
                    ws.cell(row = countRow, column = countColumn).fill = PatternFill(start_color = 'FBFBE2', fill_type = 'solid')
                    ws.cell(row = countRow, column = countColumn).font = Font(name = 'Arial', size = '11')
                    if i == 'id_tipo_doc':
                        valueRow = q.id_tipo_doc.__str__()
                    elif i == 'id_ficha':
                        valueRow = q.id_ficha.__str__()
                    elif i == 'id_rol':
                        valueRow = q.id_rol.__str__()
                    elif i == 'registro':
                        valueRow = str(getattr(q, i))
                    elif i == 'last_login':
                        valueRow = str(getattr(q, i)).split('+')[0]
                        if str(valueRow) == 'None':
                            valueRow = 'No ha ingresado al aplicativo'
                    else:
                        valueRow = getattr(q, i)
                    ws.cell(row=countRow, column=countColumn).value = valueRow
                    countRow += 1
                countColumn += 1

            # Nombre del archivo
            nombreArchivo = f'REPORTE {self.model.__name__.upper()}.xlsx'
            # Definir el tipo de respuesta
            response = HttpResponse(content_type = 'application/ms-excel')
            contenido = "attachment; filename = {0}".format(nombreArchivo)
            response['Content-Disposition'] = contenido
            wb.save(response)
            return response
        return render(request, self.template_name, {'usuarios': self.get_queryset()})

class crearUsuario(CreateView):
    model = Usuario
    template_name = 'usuarios/crear.html'
    form_class = UsuarioForm
    success_url = reverse_lazy('usuarios')

    def post(self, request, *args, **kwargs):
        if is_ajax(request=request):
            form = self.form_class(request.POST, request.FILES)
            if form.is_valid():
                form.save()
                mensaje = f'{self.model.__name__} registrado correctamente!'
                error = 'no hay error'
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 201
                return response
            else:
                mensaje = f'{self.model.__name__} no se pudo registrar'
                error = form.errors
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 400
                return response
        else:
            return redirect('usuarios')

class EditarUsuario(UpdateView):
    model = Usuario
    template_name = 'usuarios/editar.html'
    form_class = UsuarioForm2
    if User.is_staff:
        success_url = reverse_lazy('usuarios')
    else:
        success_url = reverse_lazy('interfaz')

    def post(self, request, *args, **kwargs):
        if is_ajax(request=request):
            form = self.form_class(request.POST, request.FILES, instance = self.get_object())
            if form.is_valid():
                form.save()
                mensaje = f'{self.model.__name__} actualizado correctamente!'
                error = 'no hay error'
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 201
                return response
            else:
                mensaje = f'{self.model.__name__} no se pudo actualizar'
                error = form.errors
                response = JsonResponse({'mensaje': mensaje, 'error': error})
                response.status_code = 400
                return response
        else:
            redirect('usuarios')

def activarUsuario(request, id):
    usuarioFiltrado = Usuario.objects.get(id=id)
    if request.POST:
        if usuarioFiltrado.is_active:
            usuarioFiltrado.is_active = False
            messages.warning(request, 'Se desactivó el usuario correctamente')
        else:
            usuarioFiltrado.is_active = True
            messages.success(request, 'Se activó el usuario correctamente!')
        usuarioFiltrado.save()
        return redirect('usuarios')
    return render(request, 'usuarios/activar_usuario.html', {'object': usuarioFiltrado})

class confirmarEliminarUsuario(DeleteView):
    model = Usuario
    template_name = 'usuarios/usuarios_confirm_delete.html'
    success_url = reverse_lazy('usuarios')

    def post(self, request, *args, **kwargs):
        return render(request, self.template_name)

def eliminarUsuario(request, id):
    eliminar = Usuario.objects.get(id = id)
    eliminar.delete()
    return redirect('usuarios')
# ! Modulo de usuario